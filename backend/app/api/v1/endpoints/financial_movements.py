import io
import uuid
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.dependencies import get_current_tenant, get_current_user
from app.models.tenant import Tenant
from app.models.user import User
from app.schemas.financial_movement import (
    FinancialMovementCreate,
    FinancialMovementResponse,
    FinancialMovementUpdate,
)
from app.services.financial_movement_service import FinancialMovementService

router = APIRouter(prefix="/financial-movements", tags=["Financial Movements"])


@router.post("", response_model=FinancialMovementResponse, status_code=status.HTTP_201_CREATED)
def create_financial_movement(
    payload: FinancialMovementCreate,
    db: Session = Depends(get_db),
    tenant: Tenant = Depends(get_current_tenant),
    _: User = Depends(get_current_user),
):
    service = FinancialMovementService(db)
    return service.create(tenant.id, payload)


@router.get("", response_model=list[FinancialMovementResponse])
def list_financial_movements(
    kind: str | None = Query(default=None),
    status_value: str | None = Query(default=None, alias="status"),
    source_type: str | None = Query(default=None),
    category: str | None = Query(default=None),
    third_party_name: str | None = Query(default=None),
    business_area: str | None = Query(default=None),
    needs_review: bool | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=500),
    db: Session = Depends(get_db),
    tenant: Tenant = Depends(get_current_tenant),
    _: User = Depends(get_current_user),
):
    service = FinancialMovementService(db)
    return service.list_by_tenant(
        tenant.id,
        kind=kind,
        status=status_value,
        source_type=source_type,
        category=category,
        third_party_name=third_party_name,
        business_area=business_area,
        needs_review=needs_review,
        date_from=date_from,
        date_to=date_to,
        skip=skip,
        limit=limit,
    )


@router.get("/export")
def export_financial_movements(
    kind: str | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    category: str | None = Query(default=None),
    db: Session = Depends(get_db),
    tenant: Tenant = Depends(get_current_tenant),
    _: User = Depends(get_current_user),
):
    service = FinancialMovementService(db)
    movements = service.list_by_tenant(
        tenant.id,
        kind=kind,
        date_from=date_from,
        date_to=date_to,
        category=category,
        limit=5000,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    headers = ["Fecha", "Tipo", "Tercero", "Concepto", "Categoría",
               "Base (€)", "IVA (€)", "Retención (€)", "Total (€)", "Estado", "Origen"]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, m in enumerate(movements, 2):
        ws.cell(row=row, column=1, value=str(m.movement_date) if m.movement_date else "")
        ws.cell(row=row, column=2, value="Ingreso" if m.kind == "income" else "Gasto")
        ws.cell(row=row, column=3, value=m.third_party_name or "")
        ws.cell(row=row, column=4, value=m.concept or "")
        ws.cell(row=row, column=5, value=m.category or "")
        ws.cell(row=row, column=6, value=float(m.net_amount or 0))
        ws.cell(row=row, column=7, value=float(m.tax_amount or 0))
        ws.cell(row=row, column=8, value=float(m.withholding_amount or 0))
        ws.cell(row=row, column=9, value=float(m.total_amount or 0))
        ws.cell(row=row, column=10, value=m.status or "")
        ws.cell(row=row, column=11, value=m.source_type or "")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"movimientos_{tenant.name}_{date.today()}.xlsx".replace(" ", "_")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/review-inbox", response_model=list[FinancialMovementResponse])
def get_review_inbox(
    confidence_level: str | None = Query(default=None, description="Filter by level: low, medium"),
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=500),
    db: Session = Depends(get_db),
    tenant: Tenant = Depends(get_current_tenant),
    _: User = Depends(get_current_user),
):
    """
    Returns movements flagged for review (needs_review=True),
    ordered from lowest to highest confidence.
    """
    service = FinancialMovementService(db)
    return service.list_for_review(
        tenant.id,
        confidence_level=confidence_level,
        skip=skip,
        limit=limit,
    )


# ── Parameterized routes come AFTER fixed-path routes ────────────────────────

@router.get("/{movement_id}", response_model=FinancialMovementResponse)
def get_financial_movement(
    movement_id: uuid.UUID,
    db: Session = Depends(get_db),
    tenant: Tenant = Depends(get_current_tenant),
    _: User = Depends(get_current_user),
):
    service = FinancialMovementService(db)
    movement = service.get_by_id(tenant.id, movement_id)

    if not movement:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado.")

    return movement


@router.patch("/{movement_id}", response_model=FinancialMovementResponse)
def update_financial_movement(
    movement_id: uuid.UUID,
    payload: FinancialMovementUpdate,
    db: Session = Depends(get_db),
    tenant: Tenant = Depends(get_current_tenant),
    _: User = Depends(get_current_user),
):
    service = FinancialMovementService(db)
    movement = service.update(tenant.id, movement_id, payload)

    if not movement:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado.")

    return movement


@router.delete("/{movement_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_financial_movement(
    movement_id: uuid.UUID,
    db: Session = Depends(get_db),
    tenant: Tenant = Depends(get_current_tenant),
    _: User = Depends(get_current_user),
):
    service = FinancialMovementService(db)
    deleted = service.delete(tenant.id, movement_id)

    if not deleted:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado.")

    return None